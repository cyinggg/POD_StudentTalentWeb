from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_from_directory
import openpyxl
from datetime import datetime, timedelta
from pytz import timezone
import calendar
import os
import pandas as pd
import calendar as cal_module

from werkzeug.utils import secure_filename
ALLOWED_EXTENSIONS = {"xlsx"}

app = Flask(__name__)

# ==========================
# SESSION CONFIGURATION
# ==========================
app.secret_key = "change_this_to_a_random_secret_in_production"
app.permanent_session_lifetime = timedelta(minutes=30)
MAX_SHIFTS = 100

# ==========================
# FILE PATHS
# ==========================
DATA_DIR = "data"
STUDENTCOACH_FILE = os.path.join(DATA_DIR, "StudentCoach.xlsx")
APPLICATIONS_FILE = os.path.join(DATA_DIR, "applications.xlsx")
NOTIFICATIONS_FILE = os.path.join(DATA_DIR, "notifications.xlsx")
COMMITTEE_FILE = os.path.join(DATA_DIR, "committee_calendar.xlsx")
SLOT_CONTROL_FILE = os.path.join(DATA_DIR, "slot_control.xlsx")

UPLOAD_MAP = {
    "studentcoach": STUDENTCOACH_FILE,
    "applications": APPLICATIONS_FILE,
    "notifications": NOTIFICATIONS_FILE,
    "committee": COMMITTEE_FILE,
    "slot_control": SLOT_CONTROL_FILE,
}

# ==========================
# CONTEXT PROCESSORS
# ==========================
@app.context_processor
def inject_current_time():
    return {'current_time': datetime.now()}

# ==========================
# HELPER FUNCTIONS
# ==========================
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def load_students_dict():
    """Load students into dict keyed by Student ID"""
    students = {}
    if not os.path.exists(STUDENTCOACH_FILE):
        return students

    wb = openpyxl.load_workbook(STUDENTCOACH_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        sid = str(row[1]).strip()
        div = str(row[9]).strip() if row[9] else ""
        students[sid] = {
            "full_name": str(row[0]).strip() if row[0] else "",
            "student_id": sid,
            "cluster": str(row[2]).strip() if row[2] else "",
            "degree": str(row[3]).strip() if row[3] else "",
            "current_year": str(row[4]).strip() if row[4] else "",
            "contact": str(row[5]).strip() if row[5] else "",
            "telegram": str(row[6]).strip() if row[6] else "",
            "can_id": str(row[7]).strip() if row[7] else "",
            "birthday": str(row[8]).strip() if row[8] else "",
            "division": div
        }
    return students

def validate_contact(student_id, contact):
    """Validate student contact number"""
    students = load_students_dict()
    student = students.get(student_id)
    if not student:
        return False
    # Remove spaces, dashes, +65 prefix etc.
    input_contact = ''.join(filter(str.isdigit, contact))
    stored_contact = ''.join(filter(str.isdigit, student.get("contact", "")))
    return input_contact == stored_contact

# Normalize contact number
# def normalize_contact(v):
#     return "".join(filter(str.isdigit, str(v)))
def normalize_contact(v):
    return str(v).strip()

def generate_month_slots(year: int, month: int):
    """Ensure slot_control.xlsx has rows for all dates in the month."""
    ensure_slot_control_file()
    df = pd.read_excel(SLOT_CONTROL_FILE, dtype=str)

    # Generate dates
    _, last_day = cal_module.monthrange(year, month)
    dates = [datetime(year, month, d) for d in range(1, last_day+1)]

    shift_types = ["Morning", "Afternoon", "Night"]
    slot_levels = ["L3", "L4", "L6"]
    slot_numbers = [1, 2]

    new_rows = []
    for date_obj in dates:
        month_str = date_obj.strftime("%Y-%m")
        date_str = date_obj.strftime("%Y-%m-%d")
        for shift in shift_types:
            for level in slot_levels:
                for slot in slot_numbers:
                    exists = (
                        (df['Date'] == date_str) &
                        (df['ShiftType'] == shift) &
                        (df['SlotLevel'] == level) &
                        (df['SlotNumber'] == str(slot))
                    ).any()
                    if not exists:
                        new_rows.append({
                            "Month": month_str,
                            "Date": date_str,
                            "ShiftType": shift,
                            "SlotLevel": level,
                            "SlotNumber": str(slot),
                            "IsOpen": "Closed",
                            "UpdatedBy": "",
                            "UpdatedAt": "",
                            "Remark": ""
                        })
    if new_rows:
        df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
        df.to_excel(SLOT_CONTROL_FILE, index=False)

def ensure_applications_file():
    # Ensure applications.xlsx exists
    if os.path.exists(APPLICATIONS_FILE):
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "Timestamp", "Division", "StudentID", "StudentName",
        "Date", "ShiftType", "SlotLevel", "SlotNumber",
        "Preference", "Type", "Details",
        "Status", "WaitingCount",
        "AdminDecision", "AdminRemark", "DecisionTimestamp"
    ])
    wb.save(APPLICATIONS_FILE)

def ensure_committee_file():
    if os.path.exists(COMMITTEE_FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "Timestamp","Division","StudentID","StudentName","Date",
        "Type","Details","Status"
    ])
    wb.save(COMMITTEE_FILE)

def append_application(record):
    ensure_applications_file()
    wb = openpyxl.load_workbook(APPLICATIONS_FILE)
    ws = wb.active
    ws.append(record)
    wb.save(APPLICATIONS_FILE)

def append_committee_entry(record):
    ensure_committee_file()
    wb = openpyxl.load_workbook(COMMITTEE_FILE)
    ws = wb.active
    ws.append(record)
    wb.save(COMMITTEE_FILE)

def update_application_status(date, shift, level, slot, student_id, status):
    if status not in {"Approved", "Rejected"}:
        raise ValueError(f"Invalid status: {status}")
    ensure_applications_file()
    wb = openpyxl.load_workbook(APPLICATIONS_FILE)
    ws = wb.active
    found = False
    approved_exists = False
    # First pass: check if Approved already exists
    for row in ws.iter_rows(min_row=2):
        if (
            row[1].value == "Student Coach" and
            row[4].value == date and
            row[5].value == shift and
            row[6].value == level and
            row[7].value == slot
        ):
            if str(row[11].value).strip() == "Approved":
                approved_exists = True
    # Block duplicate approvals
    if status == "Approved" and approved_exists:
        return False
    # Second pass: update target row
    for row in ws.iter_rows(min_row=2):
        if (
            row[1].value == "Student Coach" and
            row[2].value == student_id and
            row[4].value == date and
            row[5].value == shift and
            row[6].value == level and
            row[7].value == slot
        ):
            row[11].value = status
            found = True
            break
    if found:
        wb.save(APPLICATIONS_FILE)
    return found

def update_application_decision(
    date, shift, level, slot, student_id,
    decision, remark=""
):
    ensure_applications_file()
    wb = openpyxl.load_workbook(APPLICATIONS_FILE)
    ws = wb.active
    found = False

    for row in ws.iter_rows(min_row=2):
        if (
            row[1].value == "Student Coach" and
            row[2].value == student_id and
            row[4].value == date and
            row[5].value == shift and
            row[6].value == level and
            row[7].value == slot
        ):
            # Status update (update_application_status logic preserved)
            if decision == "Approved":
                row[11].value = "Approved"
            elif decision == "Rejected":
                row[11].value = "Rejected"
            else:
                row[11].value = "Approved"

            # New fields
            row[13].value = decision
            row[14].value = remark
            row[15].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            found = True
            break

    if found:
        wb.save(APPLICATIONS_FILE)
    return found

def load_committee_entries(division):
    ensure_committee_file()
    wb = openpyxl.load_workbook(COMMITTEE_FILE)
    ws = wb.active
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] != division:
            continue
        entries.append({
            "timestamp": row[0],
            "division": row[1],
            "student_id": row[2],
            "student_name": row[3],
            "date": row[4],
            "type": row[5],
            "details": row[6],
            "status": row[7]
        })
    return entries

def ensure_slot_control_file():
    if os.path.exists(SLOT_CONTROL_FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "Month", "Date", "ShiftType", "SlotLevel", "SlotNumber",
        "IsOpen", "UpdatedBy", "UpdatedAt", "Remark"
    ])
    wb.save(SLOT_CONTROL_FILE)

def load_slot_control():
    data = []
    try:
        wb = openpyxl.load_workbook(SLOT_CONTROL_FILE)
        ws = wb.active
        # Columns: Month | Date | ShiftType | SlotLevel | SlotNumber | IsOpen | UpdatedBy | UpdatedAt | Remark
        for row in ws.iter_rows(min_row=2, values_only=True):
            _, date, shift_type, slot_level, slot_number, is_open, _, _, remark = row

            # Ensure date is a string like YYYY-MM-DD
            if isinstance(date, datetime):
                date_str = date.strftime("%Y-%m-%d")
            else:
                date_str = str(date)

            # Convert is_open reliably to boolean
            if isinstance(is_open, bool):
                is_open_bool = is_open
            elif isinstance(is_open, (int, float)):
                is_open_bool = bool(is_open)
            elif isinstance(is_open, str):
                is_open_bool = is_open.strip().lower() in ["true", "1", "yes", "open"]
            else:
                is_open_bool = False  # default to closed

            data.append({
                "date": date_str,
                "shift_type": shift_type,
                "slot_level": slot_level,
                "slot_number": slot_number,
                "is_open": is_open_bool,
                "remarks": remark or ""
            })
    except Exception as e:
        print("Error loading slot_control.xlsx:", e)
    return data

def get_slot_status(month, date, shift_type, slot_level, slot_number):
    """Return tuple (is_open: bool, remark: str) for calendar"""
    is_open, remark = False, ""
    if not os.path.exists(SLOT_CONTROL_FILE):
        return is_open, remark
    try:
        df = pd.read_excel(SLOT_CONTROL_FILE, dtype=str)
    except:
        return is_open, remark

    month = str(month).strip()
    date = str(date).strip()
    shift_type = str(shift_type).strip()
    slot_level = str(slot_level).strip()
    slot_number = str(slot_number).strip()

    df["Month"] = df["Month"].astype(str).str.strip()
    df["Date"] = df["Date"].astype(str).str.strip()
    df["ShiftType"] = df["ShiftType"].astype(str).str.strip()
    df["SlotLevel"] = df["SlotLevel"].astype(str).str.strip()
    df["SlotNumber"] = df["SlotNumber"].astype(str).str.strip()

    match = df[
        (df["Month"] == month) &
        (df["Date"] == date) &
        (df["ShiftType"] == shift_type) &
        (df["SlotLevel"] == slot_level) &
        (df["SlotNumber"] == slot_number)
    ]

    if match.empty:
        return is_open, remark

    row = match.iloc[-1]
    raw_open = str(row.get("IsOpen", "")).strip().lower()
    is_open = raw_open in ["1", "true", "yes", "open"]
    remark = str(row.get("Remark", "")).strip() if not pd.isna(row.get("Remark", "")) else ""
    return is_open, remark

def is_slot_open(date, shift_type, slot_level, slot_number):
    """Return True if the slot is open according to slot_control.xlsx"""
    if not os.path.exists(SLOT_CONTROL_FILE):
        return False
    try:
        df = pd.read_excel(SLOT_CONTROL_FILE, dtype=str)
    except:
        return False

    month = str(date)[:7]
    date = str(date).strip()
    shift_type = str(shift_type).strip()
    slot_level = str(slot_level).strip()
    slot_number = str(slot_number).strip()

    df["Month"] = df["Month"].astype(str).str.strip()
    df["Date"] = df["Date"].astype(str).str.strip()
    df["ShiftType"] = df["ShiftType"].astype(str).str.strip()
    df["SlotLevel"] = df["SlotLevel"].astype(str).str.strip()
    df["SlotNumber"] = df["SlotNumber"].astype(str).str.strip()

    match = df[
        (df["Month"] == month) &
        (df["Date"] == date) &
        (df["ShiftType"] == shift_type) &
        (df["SlotLevel"] == slot_level) &
        (df["SlotNumber"] == slot_number)
    ]

    if match.empty:
        return False

    raw = str(match.iloc[-1]["IsOpen"]).strip().lower()
    return raw in ["1", "true", "yes", "open"]

# ==========================
# ROUTES
# ==========================
@app.route("/")
def start_page():
    return render_template("start.html")

@app.route("/select_division")
def select_division():
    divisions = [
        "POD Staff and Administration",
        "Student Talent Committee and Ambassador",
        "Student Coach",
        "Student Technologist",
        "Student Project Engineer",
        "Student Project Executive"
    ]
    # session.clear()
    return render_template("select_division.html", divisions=divisions)

@app.route("/start_login/<path:division>")
def start_login(division):
    session.clear()
    session["division"] = division
    return render_template("login.html", division=division)

@app.route("/verify_id", methods=["POST"])
def verify_id():
    student_id = request.form["student_id"].strip()
    students = load_students_dict()
    student = students.get(student_id)

    if not student:
        return render_template(
            "login.html",
            division=session.get("division"),
            error="Invalid Student ID"
        )

    session["student_id"] = student_id
    session["student_name"] = student.get("full_name", "")

    return render_template(
        "login.html",
        division=session.get("division"),
        student_name=session.get("student_name")
    )

@app.route("/login", methods=["POST"])
def login():
    if "student_id" not in session:
        return redirect(url_for("select_division"))

    contact_input = request.form.get("contact", "").strip()

    students = load_students_dict()
    student = students.get(session["student_id"])

    if not student:
        return redirect(url_for("select_division"))

    # ---- NORMALIZE CONTACT ----
    excel_contact = "".join(filter(str.isdigit, str(student["contact"])))
    input_contact = "".join(filter(str.isdigit, contact_input))

    if excel_contact != input_contact:
        return render_template(
            "login.html",
            division=session.get("division"),
            student_name=session.get("student_name"),
            error="Invalid contact number"
        )

    # ---- SET ROLE FROM EXCEL (NOT FROM URL) ----
    role = str(student.get("division", "")).strip()
    session["role"] = role

    print("LOGIN OK → ROLE:", role)

    # ---- REDIRECT BY ROLE ----
    if role == "POD Staff and Administration":
        return redirect(url_for("admin_home"))

    if role in [
        "Student Coach",
        "Student Talent Committee and Ambassador"
    ]:
        return redirect(url_for("calendar_view"))

    return redirect(url_for("select_division"))

@app.route("/calendar")
def calendar_view():
    # -------------------------------
    # Access control
    # -------------------------------
    if "student_id" not in session:
        return redirect(url_for("select_division"))

    if session.get("role") == "POD Staff and Administration":
        return redirect(url_for("admin_home"))

    # -------------------------------
    # Determine month/year to display
    # -------------------------------
    sg = datetime.now(timezone("Asia/Singapore"))
    try:
        year = int(request.args.get("year", sg.year))
        month = int(request.args.get("month", sg.month))
    except (TypeError, ValueError):
        year = sg.year
        month = sg.month
        # HARD guard
    if month < 1 or month > 12:
        month = sg.month

    # -------------------------------
    # Ensure slot control rows exist
    # -------------------------------
    def generate_month_slots(year, month):
        if not isinstance(month, int) or not (1 <= month <= 12):
            return  # silently skip invalid month
    generate_month_slots(year, month)
    # Also ensure next month is generated if visible
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    generate_month_slots(next_year, next_month)

    # -------------------------------
    # Prepare calendar grid
    # -------------------------------
    month_name = calendar.month_name[month]
    cal = calendar.Calendar(firstweekday=0).monthdayscalendar(year, month)

    # -------------------------------
    # Role flags
    # -------------------------------
    role = session.get("role", "")
    is_pod = False
    is_committee = role == "Student Talent Committee and Ambassador"
    is_coach = role == "Student Coach"

    # -------------------------------
    # Prepare slot data for the month
    # -------------------------------
    calendar_data = []
    shift_types = ["Morning", "Afternoon", "Night"]
    slot_levels = ["L3", "L4", "L6"]
    slot_numbers = {"L3": 2, "L4": 2, "L6": 2}

    for day in range(1, calendar.monthrange(year, month)[1]+1):
        date_str = f"{year}-{month:02d}-{day:02d}"
        for shift in shift_types:
            for level in slot_levels:
                for slot_num in range(1, slot_numbers[level]+1):
                    calendar_data.append({
                        "month": f"{year}-{month:02d}",
                        "date": date_str,
                        "shift_type": shift,
                        "slot_level": level,
                        "slot_number": slot_num
                    })

    # -------------------------------
    # Render template
    # -------------------------------
    return render_template(
        "calendar.html",
        year=year,
        month=month,
        month_name=month_name,
        cal=cal,
        role=role,
        is_pod=is_pod,
        is_committee=is_committee,
        is_coach=is_coach,
        student_name=session.get("student_name"),
        calendar_data=calendar_data,
        get_slot_status=get_slot_status
    )

@app.route("/logout", methods=["GET", "POST"])
def logout():
    session.clear()
    return redirect(url_for("select_division"))

@app.route("/projecthub")
def projecthub_home():
    return render_template("projecthub_home.html")

@app.route("/projecthub/student_coach_schedule")
def projecthub_duty_calendar():

    ensure_applications_file()  # keep

    df = pd.read_excel(APPLICATIONS_FILE)

    approved_bookings = {}

    for _, row in df.iterrows():
        if str(row.get("AdminDecision", "")).strip() != "Approved":
            continue

        date_str = row.get("Date")
        if not isinstance(date_str, str):
            date_str = df.to_datetime(date_str).strftime("%Y-%m-%d")

        shift = str(row.get("ShiftType"))
        level = str(row.get("SlotLevel"))
        slot = int(row.get("SlotNumber", 1))
        student_name = str(row.get("StudentName"))

        key = (date_str, shift, level, slot)
        approved_bookings[key] = student_name

    # ✅ MONTH NAVIGATION SUPPORT
    sg = datetime.now()
    year = request.args.get("year", type=int) or sg.year
    month = request.args.get("month", type=int) or sg.month

    if month < 1 or month > 12:
        year = sg.year
        month = sg.month

    month_name = calendar.month_name[month]
    cal = calendar.Calendar(firstweekday=0).monthdayscalendar(year, month)

    levels = sorted(df["SlotLevel"].dropna().unique())

    return render_template(
        "projecthub_duty_calendar.html",
        approved_bookings=approved_bookings,
        year=year,
        month=month,
        month_name=month_name,
        cal=cal,
        levels=levels
    )

# ==========================
# STUDENT COACH APIs
# ==========================

@app.route("/api/applications")
def api_applications():
    ensure_applications_file()
    
    # Load students data for Batch/Division
    students_dict = load_students_dict()  # keyed by StudentID
    
    # Load all applications
    wb = openpyxl.load_workbook(APPLICATIONS_FILE)
    ws = wb.active
    rows = []
    
    # Prepare waiting_map: key = (Date, ShiftType, SlotLevel, SlotNumber)
    waiting_map = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] != "Student Coach":
            continue
        if r[11] != "Pending":
            continue
        key = (r[4], r[5], r[6], r[7])
        waiting_map[key] = waiting_map.get(key, 0) + 1

    # Count total approved slots per student
    total_slots_map = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] != "Student Coach":
            continue
        student_id = str(r[2])
        total_slots_map[student_id] = total_slots_map.get(student_id, 0) + 1

    # Iterate applications to build API response
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] != "Student Coach":
            continue

        student_id = str(row[2])
        student_info = students_dict.get(student_id, {})
        batch = student_info.get("current_year", "")
        total_slots = total_slots_map.get(student_id, 0)

        rows.append({
            "timestamp": row[0],
            "division": row[1],
            "student_id": student_id,
            "student_name": row[3],
            "date": row[4],
            "shift_type": row[5],
            "slot_level": row[6],
            "slot_number": row[7],
            "preference": row[8],
            "type": row[9],
            "details": row[10],
            "status": row[11],
            "waiting_count": waiting_map.get((row[4], row[5], row[6], row[7]), 0),
            "admin_decision": row[13] if len(row) > 13 else "",
            "admin_remark": row[14] if len(row) > 14 else "",
            "batch": batch,
            "total_slots": total_slots
        })

    return jsonify(rows)

@app.route("/api/submit", methods=["POST"])
def api_submit():
    if session.get("role") != "Student Coach":
        return jsonify({"status": "error", "message": "Unauthorized"}), 403
    data = request.get_json()
    date = data.get("date")
    shift_type = data.get("shift_type")
    slot_level = data.get("slot_level")
    slot_number = data.get("slot_number")
    if not all([date, shift_type, slot_level, slot_number]):
        return jsonify({"status": "error", "message": "Missing data"}), 400
    
    if not is_slot_open(date, shift_type, slot_level, slot_number):
        return jsonify({
            "status": "error",
            "message": "This slot is currently closed by admin."
        }), 400

    ensure_applications_file()
    wb = openpyxl.load_workbook(APPLICATIONS_FILE)
    ws = wb.active
    current_pref = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == "Student Coach" and row[2] == session.get("student_id") and row[4][:7] == date[:7] and row[11] in ["Pending", "Approved"]:
            current_pref = max(current_pref, int(row[8])+1)
    if current_pref > MAX_SHIFTS:
        return jsonify({"status": "error", "message": "You can only submit up to TBC shifts per month."}), 400
    ts = datetime.now(timezone("Asia/Singapore")).strftime("%Y-%m-%d %H:%M:%S")
    record = [
        ts, "Student Coach", session.get("student_id"), session.get("student_name"),
        date, shift_type, slot_level, slot_number, current_pref, "Shift", "", "Pending"
    ]
    ws.append(record)
    wb.save(APPLICATIONS_FILE)
    return jsonify({"status": "success", "message": f"Shift booked (Pending) — Preference {current_pref}", "preference": current_pref})

@app.route("/api/coach/cancel", methods=["POST"])
def api_coach_cancel():
    if "student_id" not in session:
        return jsonify({"success": False, "message": "Not logged in"}), 401

    data = request.get_json()
    if ws.max_column < 16:
        ws.cell(row=1, column=16).value = "Cancelled At"

    date = data.get("date")
    shift = data.get("shift")
    level = data.get("level")
    slot = data.get("slot")

    if not all([date, shift, level, slot]):
        return jsonify({"success": False, "message": "Missing data"}), 400

    ensure_applications_file()
    wb = openpyxl.load_workbook(APPLICATIONS_FILE)
    ws = wb.active

    cancelled = False

    for row in ws.iter_rows(min_row=2):
        if (
            str(row[2].value) == str(session["student_id"]) and
            str(row[4].value) == str(date) and
            str(row[5].value) == str(shift) and
            str(row[6].value) == str(level) and
            str(row[7].value) == str(slot) and
            str(row[11].value) in ["Pending", "Approved"]
        ):
            row[11].value = "Cancelled"
            cancelled_col = 16  # Column P
            ws.cell(row=row[0].row, column=cancelled_col).value = \
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cancelled = True
            break

    wb.save(APPLICATIONS_FILE)

    if cancelled:
        return jsonify({"success": True})
    else:
        return jsonify({"success": False, "message": "Student Coach cancel slot not found"}), 404

@app.route("/projecthub/my_bookings")
def my_bookings():
    return render_template("my_bookings.html")

# ==========================
# ADMIN ROUTES
# ==========================
@app.route("/admin_home")
def admin_home():
    if session.get("role") != "POD Staff and Administration":
        return redirect(url_for("select_division"))
    return render_template("admin_home.html")

@app.route("/admin/upload", methods=["POST"])
def admin_upload():
    if session.get("role") != "POD Staff and Administration":
        return "Unauthorized", 403

    file = request.files.get("file")
    if not file or file.filename == "":
        flash("No file selected")
        return redirect(url_for("admin_downloads_page"))

    filename = secure_filename(file.filename)

    # Only allow exact known files
    allowed = {
        os.path.basename(STUDENTCOACH_FILE): STUDENTCOACH_FILE,
        os.path.basename(APPLICATIONS_FILE): APPLICATIONS_FILE,
        os.path.basename(SLOT_CONTROL_FILE): SLOT_CONTROL_FILE,
        os.path.basename(NOTIFICATIONS_FILE): NOTIFICATIONS_FILE,
        os.path.basename(COMMITTEE_FILE): COMMITTEE_FILE,
    }

    if filename not in allowed:
        flash("Invalid file name. Upload must match an existing system file.")
        return redirect(url_for("admin_downloads_page"))

    if not filename.lower().endswith(".xlsx"):
        flash("Only .xlsx files are allowed")
        return redirect(url_for("admin_downloads_page"))

    # SLOT CONTROL STRUCTURE VALIDATION
    if filename == os.path.basename(SLOT_CONTROL_FILE):
        try:
            df_check = pd.read_excel(file)
            required_cols = {
                "Month", "Date", "ShiftType",
                "SlotLevel", "SlotNumber", "IsOpen"
            }

            if not required_cols.issubset(df_check.columns):
                flash("slot_control.xlsx has invalid or missing columns")
                return redirect(url_for("admin_downloads_page"))

            # Reset file pointer after reading
            file.seek(0)

        except Exception as e:
            flash(f"Failed to read slot_control.xlsx: {e}")
            return redirect(url_for("admin_downloads_page"))

    target_path = allowed[filename]
    os.makedirs(os.path.dirname(target_path), exist_ok=True)

    try:
        file.save(target_path)
    except Exception as e:
        flash(f"Upload failed: {e}")
        return redirect(url_for("admin_downloads_page"))

    flash(f"{filename} replaced successfully")
    return redirect(url_for("admin_downloads_page"))

# Admin-only route to download Excel files
@app.route("/admin/download/<filename>")
def admin_download(filename):
    if session.get("role") != "POD Staff and Administration":
        return "Unauthorized", 403

    allowed_files = {
        os.path.basename(STUDENTCOACH_FILE),
        os.path.basename(APPLICATIONS_FILE),
        os.path.basename(SLOT_CONTROL_FILE),
        os.path.basename(NOTIFICATIONS_FILE),
        os.path.basename(COMMITTEE_FILE),
    }

    if filename not in allowed_files:
        return "File not found", 404

    file_path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(file_path):
        return "File missing on server", 404

    download_name = f"{datetime.now():%Y%m%d_%H%M%S}_{filename}"

    return send_from_directory(
        DATA_DIR,
        filename,
        as_attachment=True,
        download_name=download_name
    )

# Admin page to list downloadable Excel files
@app.route("/admin/downloads")
def admin_downloads_page():
    if session.get("role") != "POD Staff and Administration":
        return redirect(url_for("admin_home"))

    excel_files = sorted([
        os.path.basename(STUDENTCOACH_FILE),
        os.path.basename(APPLICATIONS_FILE),
        os.path.basename(SLOT_CONTROL_FILE),
        os.path.basename(NOTIFICATIONS_FILE),
        os.path.basename(COMMITTEE_FILE)
    ])

    return render_template("admin_downloads.html", excel_files=excel_files)

@app.route("/api/slot_control")
def api_slot_control():
    return jsonify(load_slot_control())

@app.route("/projecthub/admin_slot_control", methods=["GET", "POST"])
def admin_slot_control():
    if "role" not in session or session["role"] != "POD Staff and Administration":
        return redirect(url_for("login"))

    sg = datetime.now(timezone("Asia/Singapore"))
    year = sg.year
    month = sg.month

    # ===== Handle selected month =====
    month_selected = request.args.get("month") or request.form.get("month")
    if month_selected:
        try:
            year, month = map(int, month_selected.split("-"))
        except Exception:
            pass
    selected_month = f"{year}-{month:02d}"

    # ===== Generate all slots for the month (Option B) =====
    generate_month_slots(year, month)

    # ===== Read slot_control.xlsx =====
    df = pd.read_excel(SLOT_CONTROL_FILE, dtype={"Month": str})
    df["Month"] = df["Month"].str.strip()
    df_month = df[df["Month"] == selected_month]

    # ===== Build calendar_data for template =====
    calendar_data = {}  # {date: {shift_type: {slot_level: [slot1, slot2]}}}
    for _, row in df_month.iterrows():
        date = row["Date"].strftime("%Y-%m-%d") if isinstance(row["Date"], datetime) else str(row["Date"])
        shift_type = row["ShiftType"]
        slot_level = row["SlotLevel"]
        slot_number = row["SlotNumber"]
        raw = str(row["IsOpen"]).strip().lower()
        is_open = raw in ["open", "true", "1", "yes"]
        remark = row.get("Remark", "")

        calendar_data.setdefault(date, {})
        calendar_data[date].setdefault(shift_type, {})
        calendar_data[date][shift_type].setdefault(slot_level, [])
        calendar_data[date][shift_type][slot_level].append(
            {
                "slot_number": slot_number,
                "is_open": is_open,
                "remark": remark
            }
        )

    return render_template(
        "admin_slot_control_calendar.html",
        calendar_data=calendar_data,
        selected_month=selected_month,
        month_name=calendar.month_name[month],
        year=year,
        slot_data=df_month.to_dict(orient="records")
    )

@app.route("/admin/slot-control/update", methods=["POST"])
def admin_update_slot_control():
    # -----------------------------
    # Role protection
    # -----------------------------
    if session.get("role") != "POD Staff and Administration":
        return jsonify({"success": False, "message": "Unauthorized"}), 403

    # -----------------------------
    # Read incoming data
    # -----------------------------
    data = request.json
    month = data.get("month")
    date = data.get("date")
    shift_type = data.get("shift_type")
    slot_level = data.get("slot_level")
    slot_number = data.get("slot_number")
    is_open = data.get("is_open")
    remark = data.get("remark", "")

    # Safety check
    if not all([month, date, shift_type, slot_level, slot_number]):
        return jsonify({"success": False, "message": "Missing required fields"}), 400

    # -----------------------------
    # Load or create slot_control.xlsx
    # -----------------------------
    if os.path.exists(SLOT_CONTROL_FILE):
        df = pd.read_excel(SLOT_CONTROL_FILE)
    else:
        df = pd.DataFrame(columns=[
            "Month", "Date", "ShiftType", "SlotLevel",
            "SlotNumber", "IsOpen", "UpdatedBy", "UpdatedAt", "Remark"
        ])

    # -----------------------------
    # Ensure Remark column exists and is string type
    # -----------------------------
    if "Remark" not in df.columns:
        df["Remark"] = ""
    else:
        df["Remark"] = df["Remark"].fillna("").astype(str)

    # Ensure IsOpen column exists
    if "IsOpen" not in df.columns:
        df["IsOpen"] = "Closed"
    else:
        df["IsOpen"] = df["IsOpen"].apply(lambda x: "Open" if str(x) in ["True","Open"] else "Closed")

    # Normalize SlotNumber column
    if "SlotNumber" in df.columns:
        df["SlotNumber"] = df["SlotNumber"].fillna(0).astype(int)

    # -----------------------------
    # Identify the exact slot row
    # -----------------------------
    mask = (
        (df["Month"] == str(month)) &
        (df["Date"] == str(date)) &
        (df["ShiftType"] == str(shift_type)) &
        (df["SlotLevel"] == str(slot_level)) &
        (df["SlotNumber"] == int(slot_number))
    )

    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    updated_by = session.get("student_name", "POD Staff and Administration")

    # -----------------------------
    # Update existing row or append new
    # -----------------------------
    if mask.any():
        df.loc[mask, "IsOpen"] = "Open" if is_open else "Closed"
        df.loc[mask, "Remark"] = str(remark)
        df.loc[mask, "UpdatedBy"] = updated_by
        df.loc[mask, "UpdatedAt"] = updated_at
        action_msg = "updated"
    else:
        new_row = {
            "Month": month,
            "Date": date,
            "ShiftType": shift_type,
            "SlotLevel": slot_level,
            "SlotNumber": int(slot_number),
            "IsOpen": "Open" if is_open else "Closed",
            "UpdatedBy": updated_by,
            "UpdatedAt": updated_at,
            "Remark": str(remark)
        }
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        action_msg = "added"

    # -----------------------------
    # Save back to Excel
    # -----------------------------
    df.to_excel(SLOT_CONTROL_FILE, index=False)

    # -----------------------------
    # Return JSON with notification message
    # -----------------------------
    return jsonify({
        "success": True,
        "message": f"Slot {slot_number} ({shift_type} - {slot_level}) {action_msg} successfully."
    })

@app.route("/admin_approvals")
def admin_approvals():
    if session.get("role") != "POD Staff and Administration":
        return redirect(url_for("select_division"))
    return render_template("admin_approvals.html")

@app.route("/api/admin/pending_applications")
def admin_pending_applications():
    if session.get("role") != "POD Staff and Administration":
        return jsonify({"status": "error", "message": "Unauthorized"}), 403
    ensure_applications_file()
    wb = openpyxl.load_workbook(APPLICATIONS_FILE)
    ws = wb.active
    applications = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] != "Student Coach":
            continue
        applications.append({
            "timestamp": row[0],
            "division": row[1],
            "student_id": row[2],
            "student_name": row[3],
            "date": row[4],
            "shift_type": row[5],
            "slot_level": row[6],
            "slot_number": row[7],
            "preference": row[8],
            "type": row[9],
            "details": row[10],
            "status": row[11],
            "waiting_count": row[12],
            "admin_decision": row[13] if len(row) > 13 else "",
            "admin_remark": row[14] if len(row) > 14 else ""
        })
    return jsonify(applications)

@app.route("/api/admin/approve", methods=["POST"])
def admin_approve():
    if session.get("role") != "POD Staff and Administration":
        return jsonify({"status": "error", "message": "Unauthorized"}), 403
    data = request.get_json()
    success = update_application_status(
        date=data["date"],
        shift=data["shift"],
        level=data["level"],
        slot=data["slot"],
        student_id=data["student_id"],
        status="Approved"
    )
    return jsonify({"status": "ok" if success else "not_found"})

@app.route("/api/admin/reject", methods=["POST"])
def admin_reject():
    if session.get("role") != "POD Staff and Administration":
        return jsonify({"status": "error", "message": "Unauthorized"}), 403
    data = request.get_json()
    success = update_application_status(
        date=data["date"],
        shift=data["shift"],
        level=data["level"],
        slot=data["slot"],
        student_id=data["student_id"],
        status="Rejected"
    )
    return jsonify({"status": "ok" if success else "not_found"})

@app.route("/api/admin/decide", methods=["POST"])
def admin_decide():
    if session.get("role") != "POD Staff and Administration":
        return jsonify({"status":"error", "message":"Unauthorized"}), 403

    data = request.get_json()
    date = data.get("date")
    shift = data.get("shift")
    level = data.get("level")
    slot = data.get("slot")
    student_id = data.get("student_id")
    # Use 'status' if sent by JS, fallback to 'decision'
    decision = data.get("status") or data.get("decision")
    remark = data.get("remark", "")

    ensure_applications_file()
    wb = openpyxl.load_workbook(APPLICATIONS_FILE)
    ws = wb.active

    found = False
    for row in ws.iter_rows(min_row=2):
        if (
            row[1].value == "Student Coach" and
            str(row[2].value) == str(student_id) and
            str(row[4].value) == str(date) and
            row[5].value == shift and
            row[6].value == level and
            str(row[7].value) == str(slot)
        ):
            # Update AdminDecision & Remark
            row[13].value = decision  # AdminDecision column
            row[14].value = remark    # AdminRemark column

            # Update Status only if Approved / Rejected (quick action effect)
            if decision in ["Approved", "Rejected"]:
                row[11].value = decision # Status column
            found = True
            break

    if found:
        wb.save(APPLICATIONS_FILE)
        return jsonify({"status":"ok"})
    else:
        return jsonify({"status":"not_found", "message":"Row not found"})

def admin_reallocate():
    if session.get("role") != "POD Staff and Administration":
        return jsonify({"status":"error","message":"Unauthorized"}),403

    data = request.get_json()
    date = data.get("date")
    shift = data.get("shift")
    level = data.get("level")
    slot = data.get("slot")
    student_id = data.get("student_id")

    # Update the booking to Approved even if there is already an approved booking for the same slot
    ensure_applications_file()
    wb = openpyxl.load_workbook(APPLICATIONS_FILE)
    ws = wb.active

    found = False
    for row in ws.iter_rows(min_row=2):
        if (
            row[1].value == "Student Coach" and
            row[2].value == student_id and
            row[4].value == date and
            row[5].value == shift and
            row[6].value == level and
            row[7].value == slot
        ):
            row[11].value = "Approved"  # Status
            row[13].value = "Approved via Reallocate"  # Admin Decision or Remark column
            found = True
            break
    if found:
        wb.save(APPLICATIONS_FILE)
        return jsonify({"status":"ok"})
    return jsonify({"status":"not_found"})

# ==========================
# Flask app for Replit
# ==========================
# Health Check for UptimeRobot
@app.route("/health")
def health():
    return "UptimeRobot ok."

# ==========================
# RUN
# ==========================
if __name__ == "__main__":
    ensure_applications_file()
    ensure_committee_file()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
